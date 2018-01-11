import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill, PatternFill, Border, Side, NamedStyle, Font
from openpyxl.cell import Cell
from openpyxl.styles import colors
from xlsxwriter.utility import xl_rowcol_to_cell
import xlrd
import pandas as pd
import numpy as np
import os

class grau_excel:
    def __init__(self, workbook_path='', sheet_name=''):
        self.workbook_path = workbook_path
        self.openpy_workbook = openpyxl.load_workbook(self.workbook_path)

        if sheet_name == '':
            self.openpy_sheet = self.openpy_workbook.get_active_sheet()
        else:
            self.openpy_sheet = self.openpy_workbook.get_sheet_by_name(sheet_name)

    def value_input(self, lin, column, value):
        self.openpy_sheet.cell(row=lin, column=column).value = value

        self.openpy_workbook.save(self.workbook_path)

    def decimal_brasileiro(self, lin=2, column=1):
        '''
        :param lin: linha inicial
        :param column: coluna a ser iterada
        :return:
        '''
        for self.row in self.openpy_sheet.iter_rows():

            self.openpy_sheet.cell(row=lin, column=column).value = str(self.openpy_sheet.cell(row=lin, column=column).value)\
                .replace('.', ',')


            lin= lin + 1

        lin = 2
        for self.row in self.openpy_sheet.iter_rows():
            if self.openpy_sheet.cell(row=lin, column=column).value == 'None':
                self.openpy_sheet.cell(row=lin, column=column).value = ''
            lin = lin + 1

        self.openpy_workbook.save(self.workbook_path)


    def resize_column(self, column='', size=40):
        '''
        :param column: nome da coluna. Precisa ser em letras
        :param size: tamanho
        :return:
        '''
        if size != '':
            self.openpy_sheet.column_dimensions[column].width = size
        else:
            self.openpy_sheet.column_dimensions[column].auto_size = True

        self.openpy_workbook.save(self.workbook_path)


    def font(self, lin=1, column=1,name='Calibri', bold=False, italic=False, color=''):
        '''
        :return:
        '''
        if color == 'azul_grau':
            color = '003366'
        elif color == 'branco':
            color = "FFFFFF"

        if color == '':
            for self.row in self.openpy_sheet.iter_rows():
                self.openpy_sheet.cell(row=lin, column=column).font \
                    = Font(name=name, bold=bold, italic=italic)
                lin = lin + 1
        else:
            for self.row in self.openpy_sheet.iter_rows():
                self.openpy_sheet.cell(row=lin, column=column).font \
                    = Font(name=name, bold=bold, italic=italic, color=color)
                lin = lin + 1

        self.openpy_workbook.save(self.workbook_path)


    def header(self, header_lin=1, header_col=1, name='Calibri', bold=True, italic=False, color='branco', pattern='azul_grau'):
        '''
        :param bold:
        :param italic:
        :param color:
        :param pattern:
        :return:
        '''
        if color == 'azul_grau':
            color = '003366'
        elif color == 'branco':
            color = "FFFFFF"

        if pattern == 'azul_grau':
            pattern = '003366'
        elif pattern == 'branco':
            pattern = 'FFFFFF'

        max_col = self.openpy_sheet.max_column


        while header_col <= max_col:
            self.openpy_sheet.cell(row=header_lin, column=header_col).font \
                = Font(name=name, bold=bold, italic=italic, color=color)

            self.openpy_sheet.cell(row=header_lin, column=header_col).fill \
                = PatternFill(start_color=pattern, end_color=pattern, fill_type='solid')

            header_col = header_col + 1

        self.openpy_workbook.save(self.workbook_path)


    def border(self, lin=1, col=1, style='thin'):

        max_col = self.openpy_sheet.max_column
        max_lin = self.openpy_sheet.max_row

        thin_border = Border(top=Side(border_style=style, color='FF000000'),
                             right=Side(border_style=style, color='FF000000'),
                             bottom=Side(border_style=style, color='FF000000'),
                             left=Side(border_style=style, color='FF000000'))


        temp_lin = lin

        while col <= max_col:
            lin = temp_lin
            while lin <= max_lin:
                self.openpy_sheet.cell(row=lin, column=col).border = thin_border
                lin = lin + 1
            col = col + 1

        self.openpy_workbook.save(self.workbook_path)


    def format_number(self, column, format='float', decimals = '0'):

        max_lin = self.openpy_sheet.max_row

        lin = 2
        while lin <= max_lin:

            if format == 'percentage' or format == 'porcentagem' or format == '%':
                if decimals == '0' or decimals == 0:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '0%'
                elif decimals == '1' or decimals == 1:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '0.0%'
                elif decimals == '2' or decimals == 2:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '0.00%'
                elif decimals == '4' or decimals == 4:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '0.0000%'

            if format == 'float' or format == 'real':
                if decimals == '0' or decimals == 0:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '#,##0.00'
                elif decimals == '1' or decimals == 1:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '#,##0.00'
                elif decimals == '2' or decimals == 2:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '#,##0.00'
                elif decimals == '4' or decimals == 4:
                    self.openpy_sheet.cell(row=lin, column=column).number_format = '#,##0.00'

            if format == 'int' or format == 'real':
                self.openpy_sheet[columns].number_format = '0'

            lin = lin + 1

        self.openpy_workbook.save(self.workbook_path)
